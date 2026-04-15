"""
差旅明細整理程式
讀取 input/ 下所有 XLS 旅費彙整表，整理出：
1. 指定衛生所（佳里區/Z20）的出差明細
2. 全機關加班費彙整
3. 每人小計
輸出成 Excel 總表（每個檔案各一份 + 合併總表）
"""
import xlrd
import re
import sys
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ========== 設定 ==========
INPUT_DIR = "input"
OUTPUT_DIR = "output"
TARGET_KEYWORDS = ["佳里", "z20", "Z20"]  # 目標衛生所識別關鍵字


def open_workbook(path):
    """開啟 XLS 檔案"""
    return xlrd.open_workbook(path, formatting_info=True)


def is_target(text):
    """判斷是否為目標衛生所"""
    if not text:
        return False
    text_upper = str(text).upper().strip()
    for kw in TARGET_KEYWORDS:
        if kw.upper() in text_upper:
            return True
    return False


def parse_date(val, wb):
    """嘗試解析日期值"""
    if isinstance(val, float) and val > 40000:
        try:
            date_tuple = xlrd.xldate_as_tuple(val, wb.datemode)
            return f"{date_tuple[0]}/{date_tuple[1]}/{date_tuple[2]}"
        except:
            return str(int(val))
    return str(val).strip() if val else ""


def safe_float(val):
    """安全轉換為數值"""
    try:
        f = float(val)
        return f if f != 0 else 0
    except:
        return 0


def split_persons(person_str, total_amount):
    """
    拆分合併姓名，回傳 [(姓名, 金額), ...]
    支援格式:
      - "陳翊瑄$175+顏詩瑋$190"
      - "葉千嬅、嚴家翎$104*2人"
      - "鍾宜珍$157+侯晏筑$156"
      - "1/21葉千嬅、嚴家翎$104*2人"  (備註格式含日期)
      - "洪紅華" (單人)
    """
    if not person_str or not person_str.strip():
        return [("", total_amount)]

    text = person_str.strip()

    # 移除前面的日期 "1/21" 等
    text = re.sub(r'^\d+/\d+\s*', '', text)

    # 格式1: "姓名$金額+姓名$金額" 或 "姓名$金額、姓名$金額"
    parts = re.split(r'[+＋]', text)
    results = []

    for part in parts:
        part = part.strip()
        if not part:
            continue

        # 子分割: 用頓號分
        sub_parts = re.split(r'[、,，]', part)

        for sp in sub_parts:
            sp = sp.strip()
            if not sp:
                continue

            # 嘗試 "姓名$金額*N人" (每人金額, N人)
            m2 = re.match(r'(.+?)\$(\d+(?:\.\d+)?)\*(\d+)人?', sp)
            if m2:
                name = m2.group(1).strip()
                amt = float(m2.group(2))  # 每人金額
                # *N人 代表每人該金額，名字已在頓號分割處理
                results.append((name, amt))
            else:
                # 嘗試提取 "姓名$金額"
                m = re.match(r'(.+?)\$(\d+(?:\.\d+)?)', sp)
                if m:
                    name = m.group(1).strip()
                    amt = float(m.group(2))
                    results.append((name, amt))
                else:
                    # 純姓名無金額
                    results.append((sp, 0))

    # 如果只解析出一筆且沒有金額，用原始總金額
    if len(results) == 1 and results[0][1] == 0:
        return [(results[0][0], total_amount)]

    # 如果有些人沒金額，平均分配剩餘
    has_amount = [(n, a) for n, a in results if a > 0]
    no_amount = [(n, a) for n, a in results if a == 0]

    if no_amount and has_amount:
        used = sum(a for _, a in has_amount)
        remaining = total_amount - used
        per_person = remaining / len(no_amount) if len(no_amount) > 0 else 0
        results = has_amount + [(n, per_person) for n, _ in no_amount]
    elif not has_amount and len(results) > 1:
        # 全部沒金額，平均分配
        per_person = total_amount / len(results)
        results = [(n, per_person) for n, _ in results]

    return results if results else [("", total_amount)]


def split_record(record):
    """
    將一筆記錄中的合併姓名拆分成多筆獨立記錄
    """
    persons = split_persons(record["姓名"], record["金額"])
    if len(persons) == 1 and persons[0][0] == record["姓名"]:
        return [record]

    split_records = []
    for name, amount in persons:
        new_rec = record.copy()
        new_rec["姓名"] = name
        new_rec["金額"] = amount
        split_records.append(new_rec)
    return split_records


def get_row(sh, r):
    """取得整行資料"""
    return [sh.cell_value(r, c) for c in range(sh.ncols)]


def row_text(row):
    """整行文字"""
    return "".join(str(v) for v in row)


def is_skip_row(row):
    """判斷是否為合計/製表等跳過行"""
    t = row_text(row).replace(" ", "").replace("\u3000", "")
    return "合計" in t or "製表" in t or "股長" in t or "總計" in t


def detect_sheet_type(wb, sheet_idx):
    """
    自動偵測分頁格式類型
    回傳: 'summary', 'overtime_detail', 'overtime_summary', 'travel_detail',
          'travel_traffic', 'travel_simple', 'travel_list', 'empty'
    """
    sh = wb.sheet_by_index(sheet_idx)
    name = wb.sheet_names()[sheet_idx]

    # 空分頁
    if sh.nrows < 3:
        return 'empty'

    # 檢查是否全空
    has_data = False
    for r in range(min(sh.nrows, 10)):
        for c in range(sh.ncols):
            if str(sh.cell_value(r, c)).strip():
                has_data = True
                break
        if has_data:
            break
    if not has_data:
        return 'empty'

    # 彙總表
    if "彙總" in name or sheet_idx == 0:
        # 確認是彙總表格式
        for r in range(min(sh.nrows, 3)):
            t = row_text(get_row(sh, r))
            if "匯款明細" in t or "彙總" in t:
                return 'summary'

    # 搜尋前 10 行的關鍵字
    header_text = ""
    for r in range(min(sh.nrows, 10)):
        header_text += row_text(get_row(sh, r))

    # 類型3: 彙整表格式 (編號 | 鄉鎮別 | 金額) - 只有 3 欄
    if sh.ncols <= 4:
        for r in range(min(sh.nrows, 5)):
            row = get_row(sh, r)
            t = row_text(row)
            if "編號" in t and ("鄉鎮別" in t or "金額" in t):
                # 判斷是加班費還是其他費用
                purpose_text = ""
                for r2 in range(min(sh.nrows, 3)):
                    purpose_text += row_text(get_row(sh, r2))
                if "加班" in purpose_text or "誤餐" in purpose_text or "醫護" in purpose_text:
                    return 'overtime_summary'
                else:
                    return 'overtime_summary'  # 此格式都按彙整處理

    # 類型5: 登革熱交通/差旅格式 (姓名 | 出差日期 | 金額 | 區別)
    if "姓   名" in header_text and "出   差" in header_text:
        return 'travel_traffic'

    # 類型4/1: 衛生所 | 領受人 格式 (含差旅和加班)
    if "領受人" in header_text:
        if "加班" in name or "加班" in header_text:
            return 'overtime_detail'
        else:
            return 'travel_detail'

    # 類型7/8: 旅費明細清單 (衛生所 | 金額 | 備註)
    if "旅費明細" in header_text or "金額(元)" in header_text or "金額（元）" in header_text:
        return 'travel_list'

    # 簡易清單 (衛生所 | 金額 | 備註)
    if sh.ncols <= 5:
        for r in range(min(sh.nrows, 5)):
            row = get_row(sh, r)
            t = row_text(row)
            if "衛生所" in t and ("金額" in t or "備註" in t):
                return 'travel_list'

    return 'empty'


def extract_overtime_detail(wb, sheet_idx, file_label=""):
    """
    解析加班費明細 (衛生所 | 領受人 | 金額 | 備註)
    自動偵測金額欄位位置
    """
    sh = wb.sheet_by_index(sheet_idx)
    sheet_name = wb.sheet_names()[sheet_idx]
    records = []
    current_office = ""

    # 偵測欄位位置
    amount_col = 5  # 預設
    note_col = 7
    header_row = -1
    for r in range(min(sh.nrows, 10)):
        row = get_row(sh, r)
        t = row_text(row)
        if "領受人" in t:
            header_row = r
            for c in range(sh.ncols):
                val = str(row[c]).strip()
                if "金額" in val:
                    amount_col = c
                if "總計" in val and amount_col == c - 1:
                    pass  # 金額在總計前一欄
                if "備註" in val:
                    note_col = c
            break

    start_row = header_row + 1 if header_row >= 0 else 5

    for r in range(start_row, sh.nrows):
        row = get_row(sh, r)
        if is_skip_row(row):
            continue
        if not any(str(v).strip() for v in row):
            continue

        # 衛生所 (col 1)
        office = str(row[1]).strip() if len(row) > 1 else ""
        if office and ("衛生所" in office or "區" in office):
            if "衛生所" not in office:
                office = office + "衛生所" if "區" in office else office
            current_office = office

        # 領受人 (col 2)
        person = str(row[2]).strip() if len(row) > 2 else ""

        # 金額 - 嘗試多個欄位
        amount = 0
        if len(row) > amount_col:
            amount = safe_float(row[amount_col])
        if amount == 0 and len(row) > 5:
            amount = safe_float(row[5])
        if amount == 0 and len(row) > 4:
            amount = safe_float(row[4])

        # 備註
        note = str(row[note_col]).strip() if len(row) > note_col else ""

        if amount > 0 and current_office:
            records.append({
                "來源檔案": file_label,
                "分頁": sheet_name,
                "衛生所": current_office,
                "姓名": person,
                "日期": "",
                "金額": amount,
                "事由": note if note else sheet_name,
                "類型": "加班費"
            })

    return records


def extract_overtime_summary(wb, sheet_idx, file_label=""):
    """
    解析彙整表格式 (編號 | 鄉鎮別 | 金額)
    """
    sh = wb.sheet_by_index(sheet_idx)
    sheet_name = wb.sheet_names()[sheet_idx]
    records = []

    # 取得用途說明
    purpose = sheet_name
    for r in range(min(sh.nrows, 3)):
        t = row_text(get_row(sh, r))
        if "用途" in t:
            purpose = t.replace("用途：", "").replace("用途:", "").strip()
            break

    # 判斷類型
    fee_type = "加班費"
    if "差旅" in purpose or "旅費" in purpose:
        fee_type = "差旅費"
    elif "誤餐" in purpose:
        fee_type = "誤餐費"
    elif "醫護" in purpose:
        fee_type = "醫護費用"

    for r in range(2, sh.nrows):
        row = get_row(sh, r)
        if is_skip_row(row):
            continue

        code = str(row[0]).strip() if len(row) > 0 else ""
        office = str(row[1]).strip() if len(row) > 1 else ""
        amount = safe_float(row[2]) if len(row) > 2 else 0

        if amount > 0 and office:
            records.append({
                "來源檔案": file_label,
                "分頁": sheet_name,
                "衛生所": office,
                "姓名": "",
                "日期": "",
                "金額": amount,
                "事由": purpose,
                "類型": fee_type
            })

    return records


def extract_travel_detail(wb, sheet_idx, file_label=""):
    """
    解析差旅明細 (衛生所 | 領受人 | 金額 | 總計 | 備註)
    """
    sh = wb.sheet_by_index(sheet_idx)
    sheet_name = wb.sheet_names()[sheet_idx]
    records = []
    current_office = ""

    # 偵測欄位位置
    amount_col = 5
    note_col = 7
    header_row = -1
    for r in range(min(sh.nrows, 10)):
        row = get_row(sh, r)
        t = row_text(row)
        if "領受人" in t:
            header_row = r
            for c in range(sh.ncols):
                val = str(row[c]).strip()
                if "金額" in val:
                    amount_col = c
                if "備註" in val:
                    note_col = c
            break

    start_row = header_row + 1 if header_row >= 0 else 5

    for r in range(start_row, sh.nrows):
        row = get_row(sh, r)
        if is_skip_row(row):
            continue
        if not any(str(v).strip() for v in row):
            continue

        office = str(row[1]).strip() if len(row) > 1 else ""
        if office and ("衛生所" in office or "區" in office):
            if "衛生所" not in office:
                office = office + "衛生所" if "區" in office else office
            current_office = office

        person = str(row[2]).strip() if len(row) > 2 else ""

        amount = 0
        if len(row) > amount_col:
            amount = safe_float(row[amount_col])
        if amount == 0 and len(row) > 5:
            amount = safe_float(row[5])

        note = str(row[note_col]).strip() if len(row) > note_col else ""

        if amount > 0 and current_office:
            records.append({
                "來源檔案": file_label,
                "分頁": sheet_name,
                "衛生所": current_office,
                "姓名": person,
                "日期": "",
                "金額": amount,
                "事由": note if note else sheet_name,
                "類型": "差旅費"
            })

    return records


def extract_travel_traffic(wb, sheet_idx, file_label=""):
    """
    解析登革熱交通/差旅格式 (姓名 | 出差日期 | 金額 | 區別 | 合計 | 備註)
    """
    sh = wb.sheet_by_index(sheet_idx)
    sheet_name = wb.sheet_names()[sheet_idx]
    records = []

    # 找標題行
    start_row = 4
    for r in range(min(sh.nrows, 10)):
        row = get_row(sh, r)
        t = row_text(row)
        if "姓" in t and "名" in t:
            start_row = r + 1
            break

    for r in range(start_row, sh.nrows):
        row = get_row(sh, r)
        if is_skip_row(row):
            continue
        if not any(str(v).strip() for v in row):
            continue

        person = str(row[0]).strip() if len(row) > 0 else ""
        date_val = parse_date(row[1], wb) if len(row) > 1 else ""
        amount = safe_float(row[4]) if len(row) > 4 else 0
        district = str(row[5]).strip() if len(row) > 5 else ""
        note = str(row[7]).strip() if len(row) > 7 else ""

        if amount > 0 and person:
            office = f"{district}衛生所" if district and "衛生所" not in district else district
            records.append({
                "來源檔案": file_label,
                "分頁": sheet_name,
                "衛生所": office,
                "姓名": person,
                "日期": date_val,
                "金額": amount,
                "事由": note if note else sheet_name,
                "類型": "差旅費"
            })

    return records


def extract_travel_list(wb, sheet_idx, file_label=""):
    """
    解析旅費明細清單 (衛生所 | 金額 | 備註)
    含教育訓練格式 (XX.衛生所 | 金額 | 姓名)
    """
    sh = wb.sheet_by_index(sheet_idx)
    sheet_name = wb.sheet_names()[sheet_idx]
    records = []

    # 找標題行
    start_row = 3
    for r in range(min(sh.nrows, 5)):
        row = get_row(sh, r)
        t = row_text(row)
        if "衛生所" in t and ("金額" in t or "備註" in t):
            start_row = r + 1
            break

    for r in range(start_row, sh.nrows):
        row = get_row(sh, r)
        if is_skip_row(row):
            continue

        office_raw = str(row[0]).strip() if len(row) > 0 else ""
        amount = safe_float(row[1]) if len(row) > 1 else 0
        note = str(row[2]).strip() if len(row) > 2 else ""

        if amount > 0 and office_raw:
            # 移除前面的編號 "01." "02." 等
            office = re.sub(r'^\d+\.?\s*', '', office_raw).strip()
            if not office:
                office = office_raw

            # 從備註提取姓名
            person = ""
            if note:
                m = re.search(r'\d+/\d+\s*(.+)', note)
                if m:
                    person = m.group(1).strip()
                else:
                    person = note

            records.append({
                "來源檔案": file_label,
                "分頁": sheet_name,
                "衛生所": office,
                "姓名": person,
                "日期": "",
                "金額": amount,
                "事由": sheet_name,
                "類型": "差旅費"
            })

    return records


def extract_summary(wb, file_label=""):
    """解析彙總表 (分頁 0)"""
    sh = wb.sheet_by_index(0)
    records = []

    # 找合計欄位位置
    total_col = -1
    header_end = 3
    for r in range(min(sh.nrows, 3)):
        for c in range(sh.ncols):
            val = str(sh.cell_value(r, c)).strip()
            if val == "合計":
                total_col = c
                break

    # 動態抓取項目欄位 (col 3 到 合計欄之前)
    if total_col < 0:
        total_col = sh.ncols - 2

    headers = []
    for c in range(3, total_col):
        h1 = str(sh.cell_value(1, c)).strip() if sh.nrows > 1 else ""
        h2 = str(sh.cell_value(2, c)).strip() if sh.nrows > 2 else ""
        headers.append(h2 if h2 else h1)

    for r in range(3, sh.nrows):
        row = get_row(sh, r)
        code = str(row[1]).strip()
        office = str(row[2]).strip()
        if not code or "總" in office:
            continue
        total = safe_float(row[total_col])
        item_amounts = {}
        for i, h in enumerate(headers):
            col_idx = 3 + i
            if col_idx < len(row):
                item_amounts[h] = safe_float(row[col_idx])
        records.append({
            "來源檔案": file_label,
            "編號": code,
            "衛生所": office,
            "各項金額": item_amounts,
            "合計": total
        })

    return records, headers


# ===== Excel 輸出 =====

# 樣式定義
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
SUBTOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SUBTOTAL_FONT = Font(bold=True, size=11)
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)


def style_subtotal(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SUBTOTAL_FILL
        cell.font = SUBTOTAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)


def auto_width(ws, max_width=22):
    """自動欄寬：上限預設 22（配合直式 A4 與 wrap_text 自動換行）"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def write_sheet_target(wb_out, target_records, title_text, is_first=True):
    """佳里區明細分頁"""
    if is_first:
        ws = wb_out.active
        ws.title = "佳里區衛生所明細"
    else:
        ws = wb_out.create_sheet("佳里區衛生所明細")

    ws.merge_cells('A1:H1')
    cell = ws.cell(row=1, column=1, value=title_text)
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')

    cols = ["來源檔案", "分頁來源", "類型", "姓名", "日期", "金額", "事由", "備註"]
    for c, h in enumerate(cols, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(cols))

    row_num = 4
    subtotal = 0
    for rec in target_records:
        ws.cell(row=row_num, column=1, value=rec["來源檔案"])
        ws.cell(row=row_num, column=2, value=rec["分頁"])
        ws.cell(row=row_num, column=3, value=rec["類型"])
        ws.cell(row=row_num, column=4, value=rec["姓名"])
        ws.cell(row=row_num, column=5, value=rec["日期"])
        ws.cell(row=row_num, column=6, value=rec["金額"])
        ws.cell(row=row_num, column=6).number_format = '#,##0'
        ws.cell(row=row_num, column=7, value=rec["事由"])
        style_data(ws, row_num, len(cols))
        subtotal += rec["金額"]
        row_num += 1

    ws.cell(row=row_num, column=5, value="合計")
    ws.cell(row=row_num, column=6, value=subtotal)
    ws.cell(row=row_num, column=6).number_format = '#,##0'
    style_subtotal(ws, row_num, len(cols))
    auto_width(ws)
    return ws


def write_sheet_all(wb_out, all_records, title_text):
    """全部明細分頁"""
    ws = wb_out.create_sheet("全部明細")
    ws.merge_cells('A1:H1')
    cell = ws.cell(row=1, column=1, value=title_text)
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')

    cols = ["來源檔案", "分頁來源", "衛生所", "類型", "姓名", "日期", "金額", "事由"]
    for c, h in enumerate(cols, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(cols))

    row_num = 4
    for rec in all_records:
        ws.cell(row=row_num, column=1, value=rec["來源檔案"])
        ws.cell(row=row_num, column=2, value=rec["分頁"])
        ws.cell(row=row_num, column=3, value=rec["衛生所"])
        ws.cell(row=row_num, column=4, value=rec["類型"])
        ws.cell(row=row_num, column=5, value=rec["姓名"])
        ws.cell(row=row_num, column=6, value=rec["日期"])
        ws.cell(row=row_num, column=7, value=rec["金額"])
        ws.cell(row=row_num, column=7).number_format = '#,##0'
        ws.cell(row=row_num, column=8, value=rec["事由"])
        style_data(ws, row_num, len(cols))
        row_num += 1

    total_all = sum(r["金額"] for r in all_records)
    ws.cell(row=row_num, column=6, value="總計")
    ws.cell(row=row_num, column=7, value=total_all)
    ws.cell(row=row_num, column=7).number_format = '#,##0'
    style_subtotal(ws, row_num, len(cols))
    auto_width(ws)
    return ws


def write_sheet_person(wb_out, person_subtotals):
    """個人小計分頁"""
    ws = wb_out.create_sheet("個人小計")
    ws.merge_cells('A1:F1')
    cell = ws.cell(row=1, column=1, value="各人員費用小計")
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')

    cols = ["衛生所", "姓名", "差旅費", "加班費", "其他費用", "總計"]
    for c, h in enumerate(cols, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(cols))

    row_num = 4
    grand = {"差旅費": 0, "加班費": 0, "其他": 0}
    for (office, person), amounts in sorted(person_subtotals.items(), key=lambda x: x[0]):
        if not person:
            continue
        travel = amounts.get("差旅費", 0)
        overtime = amounts.get("加班費", 0)
        other = sum(v for k, v in amounts.items() if k not in ("差旅費", "加班費"))
        total = travel + overtime + other
        ws.cell(row=row_num, column=1, value=office)
        ws.cell(row=row_num, column=2, value=person)
        ws.cell(row=row_num, column=3, value=travel)
        ws.cell(row=row_num, column=4, value=overtime)
        ws.cell(row=row_num, column=5, value=other)
        ws.cell(row=row_num, column=6, value=total)
        for c in range(3, 7):
            ws.cell(row=row_num, column=c).number_format = '#,##0'
        style_data(ws, row_num, len(cols))
        grand["差旅費"] += travel
        grand["加班費"] += overtime
        grand["其他"] += other
        row_num += 1

    ws.cell(row=row_num, column=2, value="總計")
    ws.cell(row=row_num, column=3, value=grand["差旅費"])
    ws.cell(row=row_num, column=4, value=grand["加班費"])
    ws.cell(row=row_num, column=5, value=grand["其他"])
    ws.cell(row=row_num, column=6, value=sum(grand.values()))
    for c in range(3, 7):
        ws.cell(row=row_num, column=c).number_format = '#,##0'
    style_subtotal(ws, row_num, len(cols))
    auto_width(ws)
    return ws


def write_sheet_overtime(wb_out, overtime_records):
    """加班費彙整分頁"""
    ws = wb_out.create_sheet("加班費彙整")
    ws.merge_cells('A1:G1')
    cell = ws.cell(row=1, column=1, value="全機關加班費彙整")
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')

    cols = ["來源檔案", "分頁來源", "衛生所", "姓名", "金額", "事由", "備註"]
    for c, h in enumerate(cols, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(cols))

    row_num = 4
    for rec in overtime_records:
        ws.cell(row=row_num, column=1, value=rec["來源檔案"])
        ws.cell(row=row_num, column=2, value=rec["分頁"])
        ws.cell(row=row_num, column=3, value=rec["衛生所"])
        ws.cell(row=row_num, column=4, value=rec["姓名"])
        ws.cell(row=row_num, column=5, value=rec["金額"])
        ws.cell(row=row_num, column=5).number_format = '#,##0'
        ws.cell(row=row_num, column=6, value=rec["事由"])
        style_data(ws, row_num, len(cols))
        row_num += 1

    total = sum(r["金額"] for r in overtime_records)
    ws.cell(row=row_num, column=4, value="總計")
    ws.cell(row=row_num, column=5, value=total)
    ws.cell(row=row_num, column=5).number_format = '#,##0'
    style_subtotal(ws, row_num, len(cols))
    auto_width(ws)
    return ws


def write_sheet_summary(wb_out, all_summaries):
    """彙總表分頁"""
    for file_label, (summary_data, summary_headers) in all_summaries.items():
        # 取前 20 字作為 sheet 名
        sheet_title = f"彙總-{file_label[:20]}"
        ws = wb_out.create_sheet(sheet_title)

        col_count = 2 + len(summary_headers) + 1
        end_col = get_column_letter(col_count)
        ws.merge_cells(f'A1:{end_col}1')
        cell = ws.cell(row=1, column=1, value=f"{file_label}")
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')

        summary_cols = ["編號", "衛生所"] + summary_headers + ["合計"]
        for c, h in enumerate(summary_cols, 1):
            ws.cell(row=3, column=c, value=h)
        style_header(ws, 3, len(summary_cols))

        row_num = 4
        for rec in summary_data:
            ws.cell(row=row_num, column=1, value=rec["編號"])
            ws.cell(row=row_num, column=2, value=rec["衛生所"])
            for i, h in enumerate(summary_headers):
                val = rec["各項金額"].get(h, 0)
                ws.cell(row=row_num, column=3 + i, value=val)
                ws.cell(row=row_num, column=3 + i).number_format = '#,##0'
            ws.cell(row=row_num, column=3 + len(summary_headers), value=rec["合計"])
            ws.cell(row=row_num, column=3 + len(summary_headers)).number_format = '#,##0'

            if is_target(rec["衛生所"]) or is_target(rec["編號"]):
                for c in range(1, len(summary_cols) + 1):
                    ws.cell(row=row_num, column=c).fill = HIGHLIGHT_FILL

            style_data(ws, row_num, len(summary_cols))
            row_num += 1

        auto_width(ws)


def process_file(filepath, file_label, log_fn=None):
    """處理單一 XLS 檔案，回傳所有明細記錄（姓名已拆分）"""
    wb = open_workbook(filepath)
    all_records = []

    def log(msg):
        if log_fn:
            log_fn(msg)
        else:
            print(msg)

    log(f"\n📁 {file_label}")
    log(f"   共 {len(wb.sheet_names())} 個分頁")

    for idx in range(len(wb.sheet_names())):
        sheet_name = wb.sheet_names()[idx]
        sheet_type = detect_sheet_type(wb, idx)

        if sheet_type == 'empty' or sheet_type == 'summary':
            if sheet_type == 'summary':
                log(f"   [{idx:2d}] {sheet_name:<30} → 彙總表")
            continue

        recs = []
        if sheet_type == 'overtime_detail':
            recs = extract_overtime_detail(wb, idx, file_label)
        elif sheet_type == 'overtime_summary':
            recs = extract_overtime_summary(wb, idx, file_label)
        elif sheet_type == 'travel_detail':
            recs = extract_travel_detail(wb, idx, file_label)
        elif sheet_type == 'travel_traffic':
            recs = extract_travel_traffic(wb, idx, file_label)
        elif sheet_type == 'travel_list':
            recs = extract_travel_list(wb, idx, file_label)

        # 拆分合併姓名
        split_recs = []
        for rec in recs:
            split_recs.extend(split_record(rec))

        all_records.extend(split_recs)
        type_label = {
            'overtime_detail': '加班明細',
            'overtime_summary': '費用彙整',
            'travel_detail': '差旅明細',
            'travel_traffic': '交通差旅',
            'travel_list': '旅費清單',
        }.get(sheet_type, sheet_type)
        split_info = f" (原{len(recs)}→拆分{len(split_recs)})" if len(split_recs) != len(recs) else ""
        log(f"   [{idx:2d}] {sheet_name:<30} → {type_label} ({len(split_recs)} 筆{split_info})")

    return all_records, wb


def main():
    sys.stdout.reconfigure(encoding='utf-8')
    print("=" * 60)
    print("差旅明細整理程式 v2.0")
    print("支援多檔案批次處理")
    print("=" * 60)

    # 掃描 input 目錄
    xls_files = sorted(glob.glob(os.path.join(INPUT_DIR, "*.xls")))
    if not xls_files:
        print("\n❌ input/ 目錄中沒有找到 .xls 檔案")
        return

    print(f"\n📂 找到 {len(xls_files)} 個 XLS 檔案:")
    for f in xls_files:
        print(f"   • {os.path.basename(f)}")

    # 處理所有檔案
    all_records = []
    all_summaries = {}

    for filepath in xls_files:
        file_label = os.path.basename(filepath).replace('.xls', '')
        records, wb = process_file(filepath, file_label)
        all_records.extend(records)

        # 解析彙總表
        try:
            summary_data, summary_headers = extract_summary(wb, file_label)
            if summary_data:
                all_summaries[file_label] = (summary_data, summary_headers)
        except:
            pass

    print(f"\n{'=' * 60}")
    print(f"📊 全部共解析 {len(all_records)} 筆明細")

    # 篩選佳里區/Z20
    target_records = [r for r in all_records if is_target(r["衛生所"])]
    print(f"🎯 佳里區衛生所(Z20)相關: {len(target_records)} 筆")

    # 篩選加班費
    overtime_records = [r for r in all_records if r["類型"] == "加班費"]
    print(f"⏰ 全機關加班費: {len(overtime_records)} 筆")

    # 個人小計
    person_subtotals = {}
    for rec in all_records:
        key = (rec["衛生所"], rec["姓名"])
        if key not in person_subtotals:
            person_subtotals[key] = {}
        fee_type = rec["類型"]
        person_subtotals[key][fee_type] = person_subtotals[key].get(fee_type, 0) + rec["金額"]

    named_persons = len([k for k in person_subtotals if k[1]])
    print(f"👤 不重複人員: {named_persons} 人")

    # 產出合併總表
    print(f"\n📝 產出 Excel 總表...")
    output_path = os.path.join(OUTPUT_DIR, "差旅明細整理總表.xlsx")

    wb_out = Workbook()
    write_sheet_target(wb_out, target_records, "佳里區衛生所(Z20) 差旅費及加班費明細 (全檔合併)", is_first=True)
    write_sheet_all(wb_out, all_records, "115年衛生所差旅費 - 全部明細")
    write_sheet_person(wb_out, person_subtotals)
    write_sheet_overtime(wb_out, overtime_records)
    write_sheet_summary(wb_out, all_summaries)
    wb_out.save(output_path)
    print(f"✅ 已產出: {output_path}")

    # 顯示佳里區摘要
    print(f"\n{'=' * 60}")
    print("佳里區衛生所(Z20) 摘要:")
    print("=" * 60)
    target_total = sum(r["金額"] for r in target_records)
    for rec in target_records:
        src = rec["來源檔案"][:15]
        print(f"  {src:<16} {rec['分頁']:<25} {rec['姓名']:<12} {rec['金額']:>8,.0f}  {rec['事由'][:30]}")
    print(f"  {'':─<80}")
    print(f"  {'合計':<53} {target_total:>8,.0f}")


if __name__ == "__main__":
    main()
